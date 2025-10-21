#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
清掃スケジュール Excel変換ツール
画像をClaude APIで解析してExcelファイルを生成します
"""

import os
import sys
import json
import base64
import threading
import io
from pathlib import Path
from tkinter import filedialog, messagebox
import customtkinter as ctk
from PIL import Image, ImageTk
import anthropic
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import fitz  # PyMuPDF

# アプリケーション設定
APP_TITLE = "清掃スケジュール Excel変換ツール"
APP_VERSION = "1.0.0"
CONFIG_FILE = "config.json"

# CustomTkinter テーマ設定
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")


class CleaningScheduleApp(ctk.CTk):
    """メインアプリケーションクラス"""
    
    def __init__(self):
        super().__init__()
        
        # ウィンドウ設定
        self.title(APP_TITLE)
        self.geometry("900x750")
        self.resizable(True, True)
        self.minsize(700, 600)  # 最小サイズを設定
        
        # 変数初期化
        self.image_path = None
        self.image_display = None
        self.api_key = None
        self.is_pdf = False
        self.pdf_page_number = 1
        
        # 設定読み込み
        self.load_config()
        
        # UI構築
        self.create_widgets()
        
    def load_config(self):
        """設定ファイルを読み込む"""
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.api_key = config.get('claude_api_key', '')
            except Exception as e:
                print(f"設定読み込みエラー: {e}")
    
    def save_config(self):
        """設定ファイルを保存"""
        try:
            config = {'claude_api_key': self.api_key or ''}
            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"設定保存エラー: {e}")
    
    def create_widgets(self):
        """UI要素を作成"""
        
        # メインフレーム
        main_frame = ctk.CTkFrame(self)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # タイトル
        title_label = ctk.CTkLabel(
            main_frame,
            text=APP_TITLE,
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title_label.pack(pady=(0, 20))
        
        # ファイル選択ボタン
        select_btn = ctk.CTkButton(
            main_frame,
            text="ファイルを選択",
            command=self.select_image,
            font=ctk.CTkFont(size=16),
            height=40
        )
        select_btn.pack(pady=10)
        
        # 画像プレビューフレーム
        preview_frame = ctk.CTkFrame(main_frame)
        preview_frame.pack(fill="both", expand=True, pady=10)
        
        self.preview_label = ctk.CTkLabel(
            preview_frame,
            text="ファイルが選択されていません",
            font=ctk.CTkFont(size=14)
        )
        self.preview_label.pack(expand=True)
        
        # ファイル名表示
        self.filename_label = ctk.CTkLabel(
            main_frame,
            text="",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        )
        self.filename_label.pack(pady=5)
        
        # 進行状況バー
        self.progress_bar = ctk.CTkProgressBar(main_frame)
        self.progress_bar.pack(fill="x", pady=10)
        self.progress_bar.set(0)
        
        # ステータスラベル
        self.status_label = ctk.CTkLabel(
            main_frame,
            text="待機中...",
            font=ctk.CTkFont(size=12)
        )
        self.status_label.pack(pady=5)
        
        # PDFページ番号入力（初期は非表示）
        self.page_frame = ctk.CTkFrame(main_frame)
        self.page_frame.pack(pady=5)
        self.page_frame.pack_forget()  # 初期は非表示
        
        page_label = ctk.CTkLabel(
            self.page_frame,
            text="PDFページ番号:",
            font=ctk.CTkFont(size=12)
        )
        page_label.pack(side="left", padx=5)
        
        self.page_entry = ctk.CTkEntry(
            self.page_frame,
            width=100,
            placeholder_text="1"
        )
        self.page_entry.pack(side="left", padx=5)
        self.page_entry.insert(0, "1")
        
        # ボタンフレーム
        self.button_frame = ctk.CTkFrame(main_frame)
        self.button_frame.pack(pady=10)
        
        # 変換開始ボタン
        self.convert_btn = ctk.CTkButton(
            self.button_frame,
            text="Excel変換開始",
            command=self.start_conversion,
            font=ctk.CTkFont(size=16, weight="bold"),
            height=40,
            width=200,
            state="disabled"
        )
        self.convert_btn.pack(side="left", padx=5)
        
        # Claude APIキー設定ボタン
        settings_btn = ctk.CTkButton(
            self.button_frame,
            text="Claude APIキー設定",
            command=self.open_settings,
            font=ctk.CTkFont(size=14),
            height=40,
            width=200
        )
        settings_btn.pack(side="left", padx=5)
    
    def select_image(self):
        """画像ファイルまたはPDFを選択"""
        filetypes = [
            ("画像・PDFファイル", "*.jpg *.jpeg *.png *.pdf"),
            ("PDFファイル", "*.pdf"),
            ("画像ファイル", "*.jpg *.jpeg *.png"),
            ("JPEGファイル", "*.jpg *.jpeg"),
            ("PNGファイル", "*.png"),
            ("すべてのファイル", "*.*")
        ]
        
        filepath = filedialog.askopenfilename(
            title="清掃スケジュール画像またはPDFを選択",
            filetypes=filetypes,
            initialdir=str(Path.home())
        )
        
        if filepath:
            self.image_path = filepath
            
            # PDFかどうかを判定
            self.is_pdf = filepath.lower().endswith('.pdf')
            
            if self.is_pdf:
                # PDFの場合はページ番号入力欄を表示（ボタンフレームの前に挿入）
                self.page_frame.pack(before=self.button_frame, pady=5)
                self.display_pdf_preview(filepath)
                self.filename_label.configure(text=f"PDF: {os.path.basename(filepath)}")
                self.status_label.configure(text="PDFを選択しました（ページ番号を指定してください）")
            else:
                # 画像の場合はページ番号入力欄を非表示
                self.page_frame.pack_forget()
                self.display_image(filepath)
                self.filename_label.configure(text=os.path.basename(filepath))
                self.status_label.configure(text="画像を選択しました")
            
            self.convert_btn.configure(state="normal")
    
    def display_image(self, filepath):
        """画像をプレビュー表示"""
        try:
            # 画像を読み込み
            img = Image.open(filepath)
            
            # プレビューサイズに調整（最大600x450）
            max_width, max_height = 600, 450
            img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
            
            # PhotoImageに変換
            photo = ImageTk.PhotoImage(img)
            
            # 表示を更新
            self.preview_label.configure(image=photo, text="")
            self.preview_label.image = photo  # 参照を保持
            
            # ウィンドウサイズを調整（画像の高さに応じて）
            self.adjust_window_size(img.height)
            
        except Exception as e:
            messagebox.showerror("エラー", f"画像の読み込みに失敗しました:\n{e}")
    
    def display_pdf_preview(self, filepath):
        """PDFの最初のページをプレビュー表示"""
        try:
            # PDFを開く
            doc = fitz.open(filepath)
            
            # ページ数を取得
            page_count = len(doc)
            
            # 最初のページを画像に変換
            page = doc[0]
            pix = page.get_pixmap(matrix=fitz.Matrix(1.5, 1.5))  # 150% ズーム
            
            # PIL Imageに変換
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            
            # プレビューサイズに調整（最大600x450）
            max_width, max_height = 600, 450
            img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
            
            # PhotoImageに変換
            photo = ImageTk.PhotoImage(img)
            
            # 表示を更新
            self.preview_label.configure(image=photo, text="")
            self.preview_label.image = photo  # 参照を保持
            
            doc.close()
            
            # ページ数情報を表示
            self.filename_label.configure(
                text=f"PDF: {os.path.basename(filepath)} （全{page_count}ページ）"
            )
            
            # ウィンドウサイズを調整（画像の高さに応じて）
            self.adjust_window_size(img.height)
            
        except Exception as e:
            messagebox.showerror("エラー", f"PDFの読み込みに失敗しました:\n{e}")
    
    def adjust_window_size(self, image_height):
        """画像の高さに応じてウィンドウサイズを調整"""
        # 必要な高さを計算
        # タイトル(60) + ボタン(60) + プレビュー(image_height+20) + ファイル名(30) + 
        # プログレスバー(40) + ステータス(30) + PDFページ(40) + ボタンフレーム(60) + 余白(100)
        required_height = 60 + 60 + (image_height + 20) + 30 + 40 + 30 + 40 + 60 + 100
        
        # 現在のウィンドウサイズを取得
        current_width = self.winfo_width()
        current_height = self.winfo_height()
        
        # 必要に応じてウィンドウサイズを拡大
        if required_height > current_height:
            # 最大サイズを画面の90%に制限
            screen_height = self.winfo_screenheight()
            max_height = int(screen_height * 0.9)
            new_height = min(required_height, max_height)
            
            # ウィンドウサイズを更新
            self.geometry(f"{current_width}x{new_height}")
            
            # UIを更新
            self.update_idletasks()
    
    def open_settings(self):
        """設定ダイアログを開く"""
        settings_window = SettingsDialog(self, self.api_key)
        self.wait_window(settings_window)
        
        # 設定が更新された場合
        if hasattr(settings_window, 'result'):
            self.api_key = settings_window.result
            self.save_config()
    
    def start_conversion(self):
        """Excel変換を開始"""
        
        # APIキーチェック
        if not self.api_key:
            messagebox.showwarning(
                "APIキー未設定",
                "Claude APIキーが設定されていません。\n設定ボタンから設定してください。"
            )
            return
        
        # 画像チェック
        if not self.image_path or not os.path.exists(self.image_path):
            messagebox.showerror("エラー", "有効な画像が選択されていません")
            return
        
        # ボタンを無効化
        self.convert_btn.configure(state="disabled")
        
        # 別スレッドで処理を実行
        thread = threading.Thread(target=self.conversion_process, daemon=True)
        thread.start()
    
    def pdf_page_to_image_base64(self, pdf_path, page_number):
        """PDFの指定ページを画像（Base64）に変換"""
        try:
            doc = fitz.open(pdf_path)
            
            # ページ番号の検証（1-indexed → 0-indexed）
            page_index = page_number - 1
            if page_index < 0 or page_index >= len(doc):
                raise ValueError(f"ページ番号が範囲外です（1〜{len(doc)}ページの範囲で指定してください）")
            
            # 指定ページを取得
            page = doc[page_index]
            
            # 高解像度で画像に変換（300dpiに相当する倍率）
            zoom = 300 / 72  # PDFは72dpi、300dpiにするには約4.17倍
            pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom))
            
            # PIL Imageに変換
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            
            # JPEGとして保存（メモリ上）
            img_byte_arr = io.BytesIO()
            img.save(img_byte_arr, format='JPEG', quality=95)
            img_byte_arr.seek(0)
            
            # Base64エンコード
            image_data = base64.standard_b64encode(img_byte_arr.read()).decode('utf-8')
            
            doc.close()
            return image_data
            
        except Exception as e:
            raise Exception(f"PDF変換エラー: {str(e)}")
    
    def conversion_process(self):
        """変換処理（別スレッド）"""
        try:
            # ステップ1: 画像読み込み
            if self.is_pdf:
                # PDFの場合
                self.update_progress(0.1, "PDFを読み込んでいます...")
                
                # ページ番号を取得
                try:
                    page_num = int(self.page_entry.get())
                except ValueError:
                    raise Exception("有効なページ番号を入力してください")
                
                # PDFを画像に変換
                image_data = self.pdf_page_to_image_base64(self.image_path, page_num)
            else:
                # 画像ファイルの場合
                self.update_progress(0.1, "画像を読み込んでいます...")
                with open(self.image_path, 'rb') as f:
                    image_data = base64.standard_b64encode(f.read()).decode('utf-8')
            
            # ステップ2: Claude APIで解析
            self.update_progress(0.3, "Claude APIで解析中...")
            table_data = self.analyze_with_claude(image_data)
            
            # ステップ3: Excel生成
            self.update_progress(0.7, "Excelファイルを生成中...")
            excel_path = self.generate_excel(table_data)
            
            # ステップ4: 完了
            self.update_progress(1.0, "完了しました！")
            
            # 完了メッセージ
            self.after(100, lambda: self.show_completion(excel_path))
            
        except Exception as e:
            error_msg = f"エラーが発生しました:\n{str(e)}"
            self.after(100, lambda: messagebox.showerror("エラー", error_msg))
            self.after(100, lambda: self.update_progress(0, "エラーが発生しました"))
        finally:
            self.after(100, lambda: self.convert_btn.configure(state="normal"))
    
    def analyze_with_claude(self, image_data):
        """Claude APIで画像を解析"""
        try:
            # Anthropic クライアントを初期化
            client = anthropic.Anthropic(
                api_key=self.api_key
            )
            
            message = client.messages.create(
                model="claude-sonnet-4-5",
                max_tokens=8000,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image",
                                "source": {
                                    "type": "base64",
                                    "media_type": "image/jpeg",
                                    "data": image_data,
                                },
                            },
                            {
                                "type": "text",
                                "text": """この画像に含まれる表データを解析して、表形式のJSON（columns と rows）で出力してください。

【重要】以下のJSON形式を必ず守ってください：
- トップレベルのキーは "title", "columns", "rows" のみ
- "columns" は列名の配列
- "rows" は各行のデータをオブジェクトの配列として表現

【必須要件】
1. 表のすべての行・列を漏らさず出力（省略禁止）
2. 説明文やコメントは一切含めず、純粋なJSONのみ出力
3. コードブロック（```json）は使用しないでください

【出力例】
{
  "title": "タイトル",
  "columns": ["列1", "列2", "列3", "列4"],
  "rows": [
    {"列1": "値1", "列2": "値2", "列3": "値3", "列4": "値4"},
    {"列1": "値5", "列2": "値6", "列3": "値7", "列4": "値8"}
  ]
}

上記の形式で、画像内のすべてのデータを含むJSONを出力してください。"""
                            }
                        ]
                    }
                ]
            )
            
            # レスポンスからJSONを抽出
            response_text = message.content[0].text
            
            # デバッグ用：レスポンスをファイルに保存
            try:
                debug_dir = Path("debug_output")
                debug_dir.mkdir(exist_ok=True)
                with open(debug_dir / 'claude_response.txt', 'w', encoding='utf-8') as f:
                    f.write(response_text)
                print(f"Claude response saved to debug_output/claude_response.txt")
            except:
                pass
            
            # JSONの抽出（```json ``` で囲まれている場合に対応）
            if "```json" in response_text:
                json_start = response_text.find("```json") + 7
                json_end = response_text.find("```", json_start)
                response_text = response_text[json_start:json_end].strip()
            elif "```" in response_text:
                json_start = response_text.find("```") + 3
                json_end = response_text.find("```", json_start)
                response_text = response_text[json_start:json_end].strip()
            
            # JSONオブジェクトのみを抽出（余分なテキストを除去）
            # 最初の { から最後の } までを抽出
            json_start_bracket = response_text.find('{')
            json_end_bracket = response_text.rfind('}')
            
            if json_start_bracket != -1 and json_end_bracket != -1:
                response_text = response_text[json_start_bracket:json_end_bracket+1]
            
            # JSONパース
            table_data = json.loads(response_text)
            
            # データ構造の検証
            if 'columns' not in table_data or 'rows' not in table_data:
                raise Exception(
                    f"不正なJSON形式：'columns'と'rows'が必要です。\n"
                    f"受け取ったキー: {list(table_data.keys())}\n"
                    f"詳細は debug_output/claude_response.txt を確認してください。"
                )
            
            if not isinstance(table_data['columns'], list) or not isinstance(table_data['rows'], list):
                raise Exception("'columns'と'rows'は配列である必要があります。")
            
            if len(table_data['rows']) == 0:
                raise Exception("データ行が0件です。画像を確認してください。")
            
            print(f"解析成功: {len(table_data['columns'])}列 x {len(table_data['rows'])}行")
            
            return table_data
            
        except Exception as e:
            raise Exception(f"Claude API解析エラー: {str(e)}")
    
    def generate_excel(self, table_data):
        """Excelファイルを生成"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "清掃スケジュール"
            
            current_row = 1
            
            # タイトル行（セル結合）
            if 'title' in table_data:
                num_cols = len(table_data.get('columns', []))
                if num_cols > 0:
                    ws['A1'] = table_data['title']
                    end_col = chr(64 + min(num_cols, 26))  # 最大Z列まで
                    ws.merge_cells(f'A1:{end_col}1')
                    ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
                    ws['A1'].fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
                    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                    current_row += 1
            
            # ヘッダー行
            columns = table_data.get('columns', table_data.get('headers', []))
            if columns:
                for col_idx, header in enumerate(columns, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=header)
                    cell.font = Font(bold=True, size=10)
                    cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                current_row += 1
            
            # データ行
            if 'rows' in table_data:
                for row_data in table_data['rows']:
                    for col_idx, header in enumerate(columns, start=1):
                        value = row_data.get(header, '')
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                    current_row += 1
            
            # 列幅を自動調整
            for column_cells in ws.columns:
                max_length = 0
                column_letter = None
                for cell in column_cells:
                    try:
                        # 結合されたセルをスキップ
                        if hasattr(cell, 'column_letter'):
                            if column_letter is None:
                                column_letter = cell.column_letter
                            if cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                    except:
                        pass
                
                # 列幅を設定
                if column_letter and max_length > 0:
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存先：選択したファイルと同じディレクトリ
            source_path = Path(self.image_path)
            output_dir = source_path.parent  # 元ファイルと同じディレクトリ
            base_name = source_path.stem  # 拡張子なしのファイル名
            
            output_path = output_dir / f"{base_name}_変換結果.xlsx"
            
            # 同名ファイルがある場合は番号を追加
            counter = 1
            while output_path.exists():
                output_path = output_dir / f"{base_name}_変換結果_{counter}.xlsx"
                counter += 1
            
            wb.save(str(output_path))
            return str(output_path)
            
        except Exception as e:
            raise Exception(f"Excel生成エラー: {str(e)}")
    
    def update_progress(self, value, status_text):
        """進行状況を更新"""
        self.after(0, lambda: self.progress_bar.set(value))
        self.after(0, lambda: self.status_label.configure(text=status_text))
    
    def show_completion(self, excel_path):
        """完了ダイアログを表示"""
        result = messagebox.showinfo(
            "変換完了",
            f"Excelファイルを作成しました！\n\n保存先:\n{excel_path}\n\nフォルダを開きますか？"
        )
        
        # フォルダを開く
        if result:
            folder_path = os.path.dirname(excel_path)
            if sys.platform == 'win32':
                os.startfile(folder_path)
            elif sys.platform == 'darwin':
                os.system(f'open "{folder_path}"')
            else:
                os.system(f'xdg-open "{folder_path}"')


class SettingsDialog(ctk.CTkToplevel):
    """設定ダイアログ"""
    
    def __init__(self, parent, current_api_key):
        super().__init__(parent)
        
        self.title("設定")
        self.geometry("500x250")
        self.resizable(False, False)
        
        # モーダルにする
        self.transient(parent)
        self.grab_set()
        
        # 中央に配置
        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_y() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")
        
        # UI構築
        self.create_widgets(current_api_key)
    
    def create_widgets(self, current_api_key):
        """設定画面のUI要素を作成"""
        
        frame = ctk.CTkFrame(self)
        frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # タイトル
        title = ctk.CTkLabel(
            frame,
            text="API設定",
            font=ctk.CTkFont(size=18, weight="bold")
        )
        title.pack(pady=(0, 20))
        
        # APIキー入力
        api_label = ctk.CTkLabel(
            frame,
            text="Claude API キー:",
            font=ctk.CTkFont(size=14)
        )
        api_label.pack(anchor="w", pady=(0, 5))
        
        self.api_entry = ctk.CTkEntry(
            frame,
            width=400,
            placeholder_text="sk-ant-api03-...",
            show="*"
        )
        self.api_entry.pack(pady=(0, 5))
        
        if current_api_key:
            self.api_entry.insert(0, current_api_key)
        
        # ヘルプテキスト
        help_text = ctk.CTkLabel(
            frame,
            text="APIキーは https://console.anthropic.com/ で取得できます",
            font=ctk.CTkFont(size=11),
            text_color="gray"
        )
        help_text.pack(pady=(0, 20))
        
        # ボタンフレーム
        button_frame = ctk.CTkFrame(frame)
        button_frame.pack(pady=10)
        
        # 保存ボタン
        save_btn = ctk.CTkButton(
            button_frame,
            text="保存",
            command=self.save_settings,
            width=100
        )
        save_btn.pack(side="left", padx=5)
        
        # キャンセルボタン
        cancel_btn = ctk.CTkButton(
            button_frame,
            text="キャンセル",
            command=self.destroy,
            width=100,
            fg_color="gray"
        )
        cancel_btn.pack(side="left", padx=5)
    
    def save_settings(self):
        """設定を保存"""
        api_key = self.api_entry.get().strip()
        
        if not api_key:
            messagebox.showwarning("警告", "APIキーを入力してください")
            return
        
        if not api_key.startswith("sk-ant-"):
            messagebox.showwarning("警告", "Claude APIキーの形式が正しくありません")
            return
        
        self.result = api_key
        messagebox.showinfo("完了", "設定を保存しました")
        self.destroy()


def main():
    """メイン関数"""
    app = CleaningScheduleApp()
    app.mainloop()


if __name__ == "__main__":
    main()

