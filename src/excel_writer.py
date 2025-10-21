"""
Excel生成モジュール

openpyxlを使ってExcelファイルを生成する
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Dict, Tuple, Optional
import logging

from .config import ExcelConfig
from .table_detect import Cell, BorderStyle

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Excelファイルを生成するクラス"""
    
    def __init__(self, config: ExcelConfig):
        """
        Args:
            config: Excel設定
        """
        self.config = config
    
    def write(
        self,
        cells: List[Cell],
        cell_texts: Dict[Tuple[int, int], Tuple[str, float]],
        border_styles: Dict[Tuple[int, int], BorderStyle],
        output_path: str
    ) -> dict:
        """
        Excelファイルを生成する
        
        Args:
            cells: セルのリスト
            cell_texts: セルのテキストと信頼度
            border_styles: 罫線スタイル
            output_path: 出力ファイルパス
            
        Returns:
            生成情報の辞書
        """
        info = {}
        
        # ワークブックを作成
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        logger.info(f"Excelファイル生成開始: {len(cells)}個のセル")
        
        # 列幅と行高を設定
        self._set_dimensions(ws, cells, info)
        
        # セルを作成
        self._write_cells(ws, cells, cell_texts, border_styles, info)
        
        # ファイルを保存
        wb.save(output_path)
        logger.info(f"Excelファイル保存完了: {output_path}")
        
        return info
    
    def _set_dimensions(self, ws, cells: List[Cell], info: dict) -> None:
        """
        列幅と行高を設定する
        
        Args:
            ws: ワークシート
            cells: セルのリスト
            info: 情報を格納する辞書
        """
        if not cells:
            return
        
        # 列ごとの幅を計算
        col_widths = {}
        for cell in cells:
            if cell.colspan == 1:  # 結合セルは除外
                col = cell.col
                width_pixels = cell.width
                width_excel = width_pixels * self.config.pixel_to_column_width * self.config.column_width_scale
                
                if col not in col_widths:
                    col_widths[col] = []
                col_widths[col].append(width_excel)
        
        # 各列の平均幅を設定
        for col, widths in col_widths.items():
            avg_width = sum(widths) / len(widths)
            # Excelの列幅の範囲は0〜255
            avg_width = max(1, min(255, avg_width))
            
            col_letter = get_column_letter(col + 1)
            ws.column_dimensions[col_letter].width = avg_width
        
        # 行ごとの高さを計算
        row_heights = {}
        for cell in cells:
            if cell.rowspan == 1:  # 結合セルは除外
                row = cell.row
                height_pixels = cell.height
                height_excel = height_pixels * self.config.pixel_to_row_height * self.config.row_height_scale
                
                if row not in row_heights:
                    row_heights[row] = []
                row_heights[row].append(height_excel)
        
        # 各行の平均高さを設定
        for row, heights in row_heights.items():
            avg_height = sum(heights) / len(heights)
            # Excelの行高の範囲は0〜409.5
            avg_height = max(1, min(409.5, avg_height))
            
            ws.row_dimensions[row + 1].height = avg_height
        
        info['columns'] = len(col_widths)
        info['rows'] = len(row_heights)
        logger.info(f"寸法設定: {info['columns']}列 × {info['rows']}行")
    
    def _write_cells(
        self,
        ws,
        cells: List[Cell],
        cell_texts: Dict[Tuple[int, int], Tuple[str, float]],
        border_styles: Dict[Tuple[int, int], BorderStyle],
        info: dict
    ) -> None:
        """
        セルを書き込む
        
        Args:
            ws: ワークシート
            cells: セルのリスト
            cell_texts: セルのテキスト
            border_styles: 罫線スタイル
            info: 情報を格納する辞書
        """
        merged_count = 0
        
        # 最初にテキストとスタイルを設定
        for cell in cells:
            excel_row = cell.row + 1  # Excelは1始まり
            excel_col = cell.col + 1
            
            # セルを取得
            excel_cell = ws.cell(row=excel_row, column=excel_col)
            
            # テキストを設定
            text, confidence = cell_texts.get((cell.row, cell.col), ("", 0.0))
            excel_cell.value = text
            
            # フォントを設定
            font = Font(
                name=self.config.default_font,
                size=self.config.default_font_size
            )
            
            # ヘッダ行は太字
            if cell.row < self.config.header_row_count:
                font = Font(
                    name=self.config.default_font,
                    size=self.config.default_font_size,
                    bold=self.config.header_bold
                )
                
                # ヘッダの背景色
                if self.config.header_fill_color:
                    excel_cell.fill = PatternFill(
                        start_color=self.config.header_fill_color,
                        end_color=self.config.header_fill_color,
                        fill_type="solid"
                    )
            
            excel_cell.font = font
            
            # 配置を設定
            alignment = Alignment(
                horizontal=self.config.default_horizontal_alignment,
                vertical=self.config.default_vertical_alignment,
                wrap_text=self.config.wrap_text
            )
            excel_cell.alignment = alignment
            
            # 罫線を設定（通常セルのみ、結合セルは後で）
            border_style = border_styles.get((cell.row, cell.col))
            if border_style and cell.rowspan == 1 and cell.colspan == 1:
                border = self._create_border(border_style)
                excel_cell.border = border
        
        # 最後にセル結合を実行
        for cell in cells:
            if cell.rowspan > 1 or cell.colspan > 1:
                excel_row = cell.row + 1
                excel_col = cell.col + 1
                end_row = excel_row + cell.rowspan - 1
                end_col = excel_col + cell.colspan - 1
                
                ws.merge_cells(
                    start_row=excel_row,
                    start_column=excel_col,
                    end_row=end_row,
                    end_column=end_col
                )
                merged_count += 1
                
                # 結合セルの罫線を設定
                border_style = border_styles.get((cell.row, cell.col))
                if border_style:
                    self._apply_border_to_merged_cell(
                        ws, excel_row, excel_col, end_row, end_col, border_style
                    )
        
        info['merged_cells'] = merged_count
        logger.info(f"セル書き込み完了: 結合セル{merged_count}個")
    
    def _create_border(self, border_style: BorderStyle) -> Border:
        """
        罫線を作成する
        
        Args:
            border_style: 罫線スタイル
            
        Returns:
            Borderオブジェクト
        """
        return Border(
            top=Side(style=border_style.top) if border_style.top else None,
            bottom=Side(style=border_style.bottom) if border_style.bottom else None,
            left=Side(style=border_style.left) if border_style.left else None,
            right=Side(style=border_style.right) if border_style.right else None
        )
    
    def _apply_border_to_merged_cell(
        self,
        ws,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
        border_style: BorderStyle
    ) -> None:
        """
        結合セルに罫線を適用する
        
        Args:
            ws: ワークシート
            start_row: 開始行
            start_col: 開始列
            end_row: 終了行
            end_col: 終了列
            border_style: 罫線スタイル
        """
        # 結合セルの外枠に罫線を設定
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                
                # 上辺
                top = Side(style=border_style.top) if row == start_row and border_style.top else None
                # 下辺
                bottom = Side(style=border_style.bottom) if row == end_row and border_style.bottom else None
                # 左辺
                left = Side(style=border_style.left) if col == start_col and border_style.left else None
                # 右辺
                right = Side(style=border_style.right) if col == end_col and border_style.right else None
                
                cell.border = Border(top=top, bottom=bottom, left=left, right=right)


def create_writer(config: Optional[ExcelConfig] = None) -> ExcelWriter:
    """
    Excelライターを作成する
    
    Args:
        config: 設定（Noneの場合はデフォルト）
        
    Returns:
        ExcelWriterインスタンス
    """
    if config is None:
        config = ExcelConfig()
    
    return ExcelWriter(config)

