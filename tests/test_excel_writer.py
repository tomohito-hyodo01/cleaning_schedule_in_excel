"""
Excel生成のテスト
"""

import unittest
import tempfile
import os
from pathlib import Path

import openpyxl

from src.config import ExcelConfig
from src.excel_writer import ExcelWriter
from src.table_detect import Cell, BorderStyle


class TestExcelWriter(unittest.TestCase):
    """ExcelWriterのテスト"""
    
    def setUp(self):
        """セットアップ"""
        self.config = ExcelConfig()
        self.writer = ExcelWriter(self.config)
        
        # 一時ファイル
        self.temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        self.temp_file.close()
    
    def tearDown(self):
        """後処理"""
        # 一時ファイルを削除
        if os.path.exists(self.temp_file.name):
            os.unlink(self.temp_file.name)
    
    def test_write_simple_table(self):
        """シンプルな表を書き込む"""
        # 2x2のセルを作成
        cells = [
            Cell(row=0, col=0, x=0, y=0, width=100, height=50),
            Cell(row=0, col=1, x=100, y=0, width=100, height=50),
            Cell(row=1, col=0, x=0, y=50, width=100, height=50),
            Cell(row=1, col=1, x=100, y=50, width=100, height=50),
        ]
        
        # テキスト
        cell_texts = {
            (0, 0): ("A1", 0.9),
            (0, 1): ("B1", 0.9),
            (1, 0): ("A2", 0.9),
            (1, 1): ("B2", 0.9),
        }
        
        # 罫線スタイル
        border_styles = {
            (0, 0): BorderStyle(),
            (0, 1): BorderStyle(),
            (1, 0): BorderStyle(),
            (1, 1): BorderStyle(),
        }
        
        # Excelを書き込む
        info = self.writer.write(cells, cell_texts, border_styles, self.temp_file.name)
        
        # ファイルが作成されたことを確認
        self.assertTrue(Path(self.temp_file.name).exists())
        
        # Excelファイルを読み込んで検証
        wb = openpyxl.load_workbook(self.temp_file.name)
        ws = wb.active
        
        self.assertEqual(ws.cell(1, 1).value, "A1")
        self.assertEqual(ws.cell(1, 2).value, "B1")
        self.assertEqual(ws.cell(2, 1).value, "A2")
        self.assertEqual(ws.cell(2, 2).value, "B2")
    
    def test_write_merged_cells(self):
        """結合セルを書き込む"""
        # 結合セルを含む表
        cells = [
            Cell(row=0, col=0, x=0, y=0, width=200, height=50, colspan=2),
            Cell(row=1, col=0, x=0, y=50, width=100, height=50),
            Cell(row=1, col=1, x=100, y=50, width=100, height=50),
        ]
        
        cell_texts = {
            (0, 0): ("Merged", 0.9),
            (1, 0): ("A2", 0.9),
            (1, 1): ("B2", 0.9),
        }
        
        border_styles = {
            (0, 0): BorderStyle(),
            (1, 0): BorderStyle(),
            (1, 1): BorderStyle(),
        }
        
        # Excelを書き込む
        info = self.writer.write(cells, cell_texts, border_styles, self.temp_file.name)
        
        # 結合セルが記録されていることを確認
        self.assertGreater(info['merged_cells'], 0)
        
        # Excelファイルを読み込んで検証
        wb = openpyxl.load_workbook(self.temp_file.name)
        ws = wb.active
        
        # 結合セルの確認
        self.assertEqual(ws.cell(1, 1).value, "Merged")
        self.assertTrue(any("A1:B1" in str(m) for m in ws.merged_cells.ranges))


if __name__ == "__main__":
    unittest.main()

